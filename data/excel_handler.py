import os
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

import sys

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

class ExcelHandler:
    def __init__(self, file_path=None):
        self.headers = ["Fecha", "Oficina", "Nombres", "C.I", "Hora", "Observaciones"]
        if file_path:
            self.file_path = os.path.abspath(file_path)
        else:
            env_path = os.environ.get("EXCEL_PATH")
            if env_path:
                self.file_path = os.path.abspath(env_path)
            else:
                self.file_path = os.path.abspath(os.path.join(BASE_DIR, "registros_cedulas.xlsx"))
            
    def set_file_path(self, path):
        self.file_path = os.path.abspath(path)

    def get_file_path(self):
        return self.file_path

    def _autofit_columns(self, ws):
        """Ajusta automáticamente el ancho de las columnas de acuerdo a la longitud del contenido."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            # Margen de 3 caracteres y ancho mínimo de 10
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    def _get_current_day_name(self):
        """Retorna el nombre del día de la semana actual en español."""
        dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        return dias[datetime.now().weekday()]

    def ensure_file_exists(self):
        """
        Comprueba si el archivo Excel existe. Si no existe, lo crea con las cabeceras.
        Retorna (success: bool, message: str)
        """
        if not os.path.exists(self.file_path):
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                dia_semana = self._get_current_day_name()
                ws.title = dia_semana
                ws.append(self.headers)
                self._autofit_columns(ws)
                
                # Proteger la hoja contra modificaciones de usuario
                ws.protection.sheet = True
                ws.protection.password = "proincoadmin2026"
                
                wb.save(self.file_path)
                wb.close()
                return True, "Archivo creado exitosamente."
            except Exception as e:
                return False, f"Error al crear el archivo Excel: {str(e)}"
        else:
            try:
                # Comprobar si se puede abrir y si tiene cabeceras
                wb = openpyxl.load_workbook(self.file_path)
                
                # Asegurar que exista la pestaña del día actual
                dia_semana = self._get_current_day_name()
                if dia_semana not in wb.sheetnames:
                    ws = wb.create_sheet(title=dia_semana)
                    ws.append(self.headers)
                    self._autofit_columns(ws)
                    ws.protection.sheet = True
                    ws.protection.password = "proincoadmin2026"
                    wb.save(self.file_path)
                
                # Si existe, verificar si la primera fila está vacía y escribir cabeceras si es necesario
                ws = wb[dia_semana]
                first_row = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
                if not any(first_row):
                    ws.append(self.headers)
                    self._autofit_columns(ws)
                    ws.protection.sheet = True
                    ws.protection.password = "proincoadmin2026"
                    wb.save(self.file_path)
                
                wb.close()
                return True, "Archivo Excel cargado exitosamente."
            except PermissionError:
                return False, "El archivo Excel está abierto en otro programa. Por favor, ciérrelo e intente de nuevo."
            except Exception as e:
                return False, f"Error al abrir el archivo Excel: {str(e)}"

    def add_record(self, nombres, ci, oficina="Matriz", observaciones=""):
        """
        Agrega un nuevo registro al final del archivo Excel en la pestaña del día de la semana actual.
        Retorna (success: bool, message: str)
        """
        # Asegurar que el directorio contenedor exista
        dir_name = os.path.dirname(self.file_path)
        if dir_name and not os.path.exists(dir_name):
            try:
                os.makedirs(dir_name, exist_ok=True)
            except Exception as e:
                return False, f"No se pudo crear la carpeta para el Excel: {str(e)}"

        # Asegurar archivo creado
        success, msg = self.ensure_file_exists()
        if not success:
            return False, msg

        try:
            wb = openpyxl.load_workbook(self.file_path)
            
            dia_semana = self._get_current_day_name()
            if dia_semana not in wb.sheetnames:
                ws = wb.create_sheet(title=dia_semana)
                ws.append(self.headers)
            else:
                ws = wb[dia_semana]
            
            now = datetime.now()
            fecha_str = now.strftime("%Y-%m-%d")
            hora_str = now.strftime("%H:%M:%S")
            
            row_data = [fecha_str, oficina, nombres, ci, hora_str, observaciones]
            ws.append(row_data)
            self._autofit_columns(ws)
            
            # Proteger la hoja contra modificaciones de usuario antes de guardar
            ws.protection.sheet = True
            ws.protection.password = "proincoadmin2026"
            
            wb.save(self.file_path)
            wb.close()
            return True, "Registro guardado correctamente."
        except PermissionError:
            return False, "Error de permisos: El archivo Excel está abierto. Ciérrelo e intente nuevamente."
        except Exception as e:
            return False, f"No se pudo guardar en Excel: {str(e)}"

